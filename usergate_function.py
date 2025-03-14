##############################################################################
####Imports lib###
from PySide2 import QtCore, QtWidgets, QtGui
import sys,os
from openpyxl import Workbook, load_workbook
import shutil
##############################################################################
###Import interface file###
from interface_ui import *
from Custom_Widgets import *
##############################################################################
###FUnction CLASS###
#Hover blok
class HoverWindow(QWidget):
    def __init__(self, text, parent=None):
        super(HoverWindow, self).__init__(parent)
        self.setWindowFlags(Qt.ToolTip)
        self.layout = QVBoxLayout()
        self.label = QLabel(text) 
        self.setFixedSize(450, 50)
        self.layout.addWidget(self.label)
        self.setLayout(self.layout)
        self.setStyleSheet("""
            color: rgb(255, 255, 255);
            background-color: rgb(13, 18, 22);
        """) 
#MAIN init  
class Usergate_creator(QMainWindow):
    def __init__(self, obj):
        QMainWindow.__init__(self)
        self.UG = obj
    #add Text coment
        self.text_coment = ""
    #add TP status
        self.TP_status = None
    #add Fstec status    
        self.fstec_status = None
    #add parameters in listView
        self.saved_structure = None
    #add parameters in TreeView
        self.selected_name_in_view = None
    #add parametrs in years
        self.value_year_tp = 1
        self.value_year_moduls = 1   
    #add parameters for block hover
        self.hover_Сard_8_1 = HoverWindow("Сетевая карта 8*1 Гб/c для UserGate D200, D500, E1000, E3000 и F8000", self)
        self.hover_Сard_8_1.hide()
        self.hover_Сard_4_1 = HoverWindow("Сетевая карта 4*1 Гб/c для UserGate D200, D500, E1000, E3000 и F8000", self)
        self.hover_Сard_4_1.hide()
        self.hover_Сard_4_1_FO = HoverWindow("Сетевая карта 4*1 Гб/с (FO) для UserGate D200, D500, E1000, E3000 и F8000", self)
        self.hover_Сard_4_1_FO.hide()
        self.hover_Сard_4_10 = HoverWindow("Сетевая карта 4*10 Гб/c для UserGate D200, D500, E1000, E3000 и F8000", self)
        self.hover_Сard_4_10.hide()
        self.hover_Сard_4_10_FO = HoverWindow("Сетевая карта 4*10 Гб/с (FO) для UserGate D200, D500, E1000, E3000 и F8000", self)
        self.hover_Сard_4_10_FO.hide()
        self.hover_Сard_1_40 = HoverWindow("Сетевая карта 1*40 Гб/с для UserGate D200, D500, E1000, E3000 и F8000", self)
        self.hover_Сard_1_40.hide()
        self.hover_SFP_RJ_45_1G = HoverWindow("Модуль медный гигабитный, 1 Гб/с, до 100 м", self)
        self.hover_SFP_RJ_45_1G.hide()
        self.hover_SFP_10G = HoverWindow("Модуль оптический (10Gb, 850nm, 500m., LC, DDM", self)
        self.hover_SFP_10G.hide()
        self.hover_SFP_10Gb_500m = HoverWindow("Модуль оптический (10Gb, SM, 500m)", self)
        self.hover_SFP_10Gb_500m.hide()
        self.hover_SFP_10Gb_10km = HoverWindow("Модуль оптический (10Gb, SM, 10km, 1310nm, LC, DDM)", self)
        self.hover_SFP_10Gb_10km.hide()
        self.hover_SFP_SX_1_25 = HoverWindow("Модуль оптический FT-SFP-SX-1,25-850-0,5-D (1,25G, 550m, Tx850nm, MM, LC, DDM)", self)
        self.hover_SFP_SX_1_25.hide()
    #Timers
        self.timer_Сard_8_1 = QTimer()
        self.timer_Сard_8_1.timeout.connect(self.WINdow_Сard_8_1)
        self.timer_Сard_4_1 = QTimer()
        self.timer_Сard_4_1.timeout.connect(self.WINdow_Сard_4_1)
        self.timer_Сard_4_1_FO = QTimer()
        self.timer_Сard_4_1_FO.timeout.connect(self.WINdow_Сard_4_1_FO)
        self.timer_Сard_4_10 = QTimer()
        self.timer_Сard_4_10.timeout.connect(self.WINdow_Сard_4_10)
        self.timer_Сard_4_10_FO = QTimer()
        self.timer_Сard_4_10_FO.timeout.connect(self.WINdow_Сard_4_10_FO)
        self.timer_Сard_1_40 = QTimer()
        self.timer_Сard_1_40.timeout.connect(self.WINdow_Сard_1_40)
        self.timer_SFP_RJ_45_1G = QTimer()
        self.timer_SFP_RJ_45_1G.timeout.connect(self.WINdow_SFP_RJ_45_1G)
        self.timer_SFP_10G = QTimer()
        self.timer_SFP_10G.timeout.connect(self.WINdow_SFP_10G)
        self.timer_SFP_10Gb_500m = QTimer()
        self.timer_SFP_10Gb_500m.timeout.connect(self.WINdow_SFP_10Gb_500m)
        self.timer_SFP_10Gb_10km = QTimer()
        self.timer_SFP_10Gb_10km.timeout.connect(self.WINdow_SFP_10Gb_10km)
        self.timer_SFP_SX_1_25 = QTimer()
        self.timer_SFP_SX_1_25.timeout.connect(self.WINdow_SFP_SX_1_25)
    #iscl
        self.not_add = ["License Security Updates", "License Advanced Threat Protection", "License Streaming antivirus", "License Mail Security", "License Management Center", "License Log Analyzer", "License Cluster", "Сard 8*1 GB/s", "Сard 4*1 GB/s", "Сard 4*1 GB/s (FO)", "Сard 4*10 GB/s", "Сard 4*10 GB/s (FO)", "Сard 1*40 GB/s", "SFP-RJ-45 1G", "SFP-RJ-45 10G", "SFP+ 10Gb, 500m", "SFP 10Gb, 10km", "SFP-SX-1,25"] 
    ##############################################################################
    ###Add mouse events no selections###
    def tree_widget_mouse_press_event(self, event):
        # Если клик был вне элемента, снимаем выделение
        self.item = self.UG.ui.treeWidget.itemAt(event.pos())
        if not self.item:
            #self.UG.ui.treeWidget.clearSelection()
            self.UG.ui.treeWidget.setSelectionMode(QtWidgets.QAbstractItemView.ContiguousSelection)
        # Вызываем оригинальный обработчик события
        QTreeWidget.mousePressEvent(self.UG.ui.treeWidget, event)
        self.selected_name_in_view = None
    ##############################################################################
    ###event roll###   
    def wheelEvent(self, event):
        # Получаем текущее значение горизонтальной прокрутки
        self.scroll_bar = self.UG.ui.scrollArea.horizontalScrollBar()
        self.current_value = self.scroll_bar.value()
        # Изменяем значение прокрутки в зависимости от направления прокрутки колесика
        if event.angleDelta().y() > 0:
            self.scroll_bar.setValue(self.current_value - 50)  # Прокрутка влево
        else:
            self.scroll_bar.setValue(self.current_value + 50)  # Прокрутка вправо
    def wheelEvent_2(self, event_2):
        self.scroll_bar_3 = self.UG.ui.scrollArea_3.horizontalScrollBar()
        self.current_value_3 = self.scroll_bar_3.value()
        if event_2.angleDelta().y() > 0:
            self.scroll_bar_3.setValue(self.current_value_3 - 50) 
        else:
            self.scroll_bar_3.setValue(self.current_value_3 + 50) 
    ##############################################################################
    ###event add_usergates### 
    def UG_add_C100(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate C100")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_C150(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate C150")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)       
    def UG_add_D200(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate D200")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_D500(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate D500")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_E1000(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate E1000")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_E3000(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate E3000")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_F8000(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate F8000")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_X1(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate X1")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_FG(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate FG")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE100(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE100")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE200(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE200")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE250(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE250")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE500(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE500")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE2000(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE2000")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE4000(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE4000")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    def UG_add_VE6000(self):
        self.parent_item = QTreeWidgetItem(self.UG.ui.treeWidget)
        self.parent_item.setText(0, "Usergate VE6000")
        self.child_item = QTreeWidgetItem(self.parent_item)
        self.child_item.setText(0, "License Security Updates")
        self.parent_item.setExpanded(True)
    ##############################################################################
    ###event Copy_file###
    def copy_file(self):
        # Получаем текст из QTextEdit
        self.new_file_name = self.UG.ui.File_name_edit.text().strip()
        # Проверяем, что имя файла не пустое
        if not self.new_file_name:
            QMessageBox.warning(self, "Ошибка", "Введите имя файла!")
            return
        else:
            #Добавляем расширение .xlsx к имени файла
            self.new_file_name += ".xlsx"
            # Путь к исходному файлу
            self.source_file = os.path.join('specifications', 'Usergate.xlsx')
            # Путь к новому файлу
            self.destination_file = self.new_file_name
            try:
                # Копируем файл
                shutil.copy(self.source_file, self.destination_file)
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось скопировать файл: {str(e)}")
    ##############################################################################
    ###event Download_scpc###
    def Download_spec(self):
        if self.UG.ui.treeWidget.topLevelItemCount() == 0:
            QMessageBox.warning(self, 'Ошибка', 'Создайте спецификацию')
            return
        # Копируем файл
        self.copy_file()
        # Открываем копию файла
        try:
            self.wb = load_workbook(self.new_file_name)
            self.ws = self.wb.worksheets[4]  # 5-я страница (индекс 4)
            # Записываем главные элементы
            for self.i in range(self.UG.ui.treeWidget.topLevelItemCount()):
                self.top_item = self.UG.ui.treeWidget.topLevelItem(self.i)
                if self.top_item.text(0).startswith('Usergate'):
                    self.ws.cell(row=3, column=3 + self.i, value=self.top_item.text(0))
                    if 'VE' in self.top_item.text(0):
                        self.ws.cell(row=6, column=3 + self.i, value="Виртуальное")
                    else:
                        self.ws.cell(row=6, column=3 + self.i, value="ПАК")
                    # Записываем дочерние элементы
                    for self.j in range(self.top_item.childCount()):
                        self.child_item = self.top_item.child(self.j)
                        if self.child_item.text(0).startswith('License'):
                            self.ws.cell(row=16 + self.j, column=3 + self.i, value=self.child_item.text(0))
                            if 'Cluster' in self.child_item.text(0):
                                self.ws.cell(row=5, column=3 + self.i, value="Да")
                                self.ws.cell(row=4, column=3 + self.i, value="2")
                            else:
                                self.ws.cell(row=5, column=3 + self.i, value="Нет")
                                self.ws.cell(row=4, column=3 + self.i, value="1")
                        elif self.child_item.text(0).startswith('Сard') or self.child_item.text(0).startswith('SFP'):
                            self.ws.cell(row=6 + self.j, column=3 + self.i, value=self.child_item.text(0))
                elif self.top_item.text(0).startswith('License'):
                    self.licenses = []
                    for self.k in range(self.UG.ui.treeWidget.topLevelItemCount()):
                        self.license_item = self.UG.ui.treeWidget.topLevelItem(self.k)
                        if self.license_item.text(0).startswith('License'):
                            self.licenses.append(self.license_item.text(0))
                    self.ws['C24'] = 'Дополнительные лицензии: ' + ', '.join(self.licenses)
                elif self.top_item.text(0).startswith('Сard') or self.top_item.text(0).startswith('SFP'):
                    self.card_sfp = []
                    for self.c in range(self.UG.ui.treeWidget.topLevelItemCount()):
                        self.card_sfp_item = self.UG.ui.treeWidget.topLevelItem(self.c)
                        if self.card_sfp_item.text(0).startswith('Сard') or self.card_sfp_item.text(0).startswith('SFP'):
                            self.card_sfp.append(self.card_sfp_item.text(0))
                    self.ws['C25'] = 'Дополнительные модули: ' + ', '.join(self.card_sfp)    
            if self.fstec_status == "Да":
                self.ws['C14'] = 'Да, ко всему оборудованию'
            elif self.fstec_status == "Нет":
                self.ws['C14'] = 'Нет'
            else:
                pass
            if self.TP_status == "Standart":
                self.ws['С23'] = f'Standart {self.value_year_tp} Год(а)\Лет'
            elif self.TP_status == "Premium":
                self.ws['C23'] = f'Premium {self.value_year_tp} Год(а)\Лет'
            else:
                self.ws['C23'] = 'Нет'
            self.ws['C15'] = f"{self.value_year_moduls} Год(а)\Лет"
            self.ws['C26'] = self.text_coment   
            # Проверяем, что имя файла не пустое
            if not self.new_file_name:
                QMessageBox.warning(self, "Ошибка", "Введите имя файла!")
                return
            else:
                self.wb.save(self.new_file_name)
            QMessageBox.information(self, 'Success', 'Файл успешно сохранен')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Ошибка при сохранении файла: {e}')
    ##############################################################################
    ###event remove string and all###
    def Remove_line(self):
        self.index = self.UG.ui.treeWidget.currentItem()
        if self.index:
            self.root = self.UG.ui.treeWidget.invisibleRootItem()
            (self.index.parent() or self.root).removeChild(self.index)
        else:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана строка для удаления')
    def Remove_all(self):
        self.UG.ui.treeWidget.clear()
    ##############################################################################
    ###event add License###
    def License_add_ATP(self):
        if self.selected_name_in_view == None:
            #Если элемент не выбран, то добавить как основной
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Advanced Threat Protection")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            # Получаем выбранный элемент (если есть)
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                #Если выбран элемент, то добавить как дочерний
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Advanced Threat Protection")
            else:
                #Если элемент не выбран, то добавить как основной
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Advanced Threat Protection")
                self.selected_name_in_view = None
    def License_add_SU(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Security Updates")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Security Updates")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Security Updates")
                self.selected_name_in_view = None
    def License_add_antivirus(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Streaming antivirus")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
            self.selected_name_in_view = None
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Streaming antivirus")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Streaming antivirus")
                self.selected_name_in_view = None
    def License_add_Mail(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Mail Security")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Mail Security")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Mail Security")
                self.selected_name_in_view = None
    def License_add_MGMT(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Management Center")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Management Center")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Management Center")
                self.selected_name_in_view = None
    def License_add_Log(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Log Analyzer")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Log Analyzer")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Log Analyzer")
                self.selected_name_in_view = None
    def License_add_Cluster(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Cluster")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "License Cluster")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "License Cluster")
                self.selected_name_in_view = None
    ##############################################################################
    ###event add card and moduls###
    def Сard_8_1(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 8*1 GB/s")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "Сard 8*1 GB/s")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 8*1 GB/s")
                self.selected_name_in_view = None
    def Сard_4_1(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*1 GB/s")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "Сard 4*1 GB/s")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*1 GB/s")
                self.selected_name_in_view = None
    def Сard_4_1_FO(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*1 GB/s (FO)")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "Сard 4*1 GB/s (FO)")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*1 GB/s (FO)")
                self.selected_name_in_view = None
    def Сard_4_10(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*10 GB/s")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "Сard 4*10 GB/s")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*10 GB/s")
                self.selected_name_in_view = None
    def Сard_4_10_FO(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*10 GB/s (FO)")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "Сard 4*10 GB/s (FO)")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 4*10 GB/s (FO)")
                self.selected_name_in_view = None
    def Сard_1_40(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 1*40 GB/s")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "Сard 1*40 GB/s")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "Сard 1*40 GB/s")
                self.selected_name_in_view = None
    def SFP_RJ_45_1G(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP-RJ-45 1G")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "SFP-RJ-45 1G")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP-RJ-45 1G")
                self.selected_name_in_view = None
    def SFP_10G(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP-RJ-45 10G")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "SFP-RJ-45 10G")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP-RJ-45 10G")
                self.selected_name_in_view = None
    def SFP_10Gb_500m(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP+ 10Gb, 500m")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "SFP+ 10Gb, 500m")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP+ 10Gb, 500m")
                self.selected_name_in_view = None
    def SFP_10Gb_10km(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP 10Gb, 10km")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "SFP 10Gb, 10km")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP 10Gb, 10km")
                self.selected_name_in_view = None
    def SFP_SX_1_25(self):
        if self.selected_name_in_view == None:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP-SX-1,25")
        elif self.selected_name_in_view in self.not_add:
            QMessageBox.warning(self, 'Ошибка', 'Нельзя добавить к выбранному элементу.')
        else:
            self.selected_item = self.UG.ui.treeWidget.currentItem()
            if self.selected_item:
                self.new_item = QTreeWidgetItem(self.selected_item)
                self.new_item.setText(0, "SFP-SX-1,25")
            else:
                self.new_item = QTreeWidgetItem(self.UG.ui.treeWidget)
                self.new_item.setText(0, "SFP-SX-1,25")
                self.selected_name_in_view = None
    def start_timer_Сard_8_1(self, event):
        self.timer_Сard_8_1.start(2000)  # 2000 milliseconds = 2 seconds
    def WINdow_Сard_8_1(self):
        self.timer_Сard_8_1.stop()
        self.hover_Сard_8_1.move(self.UG.ui.card_1_btn_1.mapToGlobal(self.UG.ui.card_1_btn_1.rect().bottomLeft()))
        self.hover_Сard_8_1.show()
    def WINdow_hide_Сard_8_1(self, event):
        self.timer_Сard_8_1.stop()
        self.hover_Сard_8_1.hide()
    def start_timer_Сard_4_1(self, event):
        self.timer_Сard_4_1.start(2000)  # 2000 milliseconds = 2 seconds
    def WINdow_Сard_4_1(self):
        self.timer_Сard_4_1.stop()
        self.hover_Сard_4_1.move(self.UG.ui.card_2_btn_2.mapToGlobal(self.UG.ui.card_2_btn_2.rect().bottomLeft()))
        self.hover_Сard_4_1.show()
    def WINdow_hide_Сard_4_1(self, event):
        self.timer_Сard_4_1.stop()
        self.hover_Сard_4_1.hide()
    def start_timer_Сard_4_1_FO(self, event):
        self.timer_Сard_4_1_FO.start(2000)  # 2000 milliseconds = 2 seconds
    def WINdow_Сard_4_1_FO(self):
        self.timer_Сard_4_1_FO.stop()
        self.hover_Сard_4_1_FO.move(self.UG.ui.card_3_btn_3.mapToGlobal(self.UG.ui.card_3_btn_3.rect().bottomLeft()))
        self.hover_Сard_4_1_FO.show()
    def WINdow_hide_Сard_4_1_FO(self, event):
        self.timer_Сard_4_1_FO.stop()
        self.hover_Сard_4_1_FO.hide()
    def start_timer_Сard_4_10(self, event):
        self.timer_Сard_4_10.start(2000)  # 2000 milliseconds = 2 seconds
    def WINdow_Сard_4_10(self):
        self.timer_Сard_4_10.stop()
        self.hover_Сard_4_10.move(self.UG.ui.card_4_btn_4.mapToGlobal(self.UG.ui.card_4_btn_4.rect().bottomLeft()))
        self.hover_Сard_4_10.show()
    def WINdow_hide_Сard_4_10(self, event):
        self.timer_Сard_4_10.stop()
        self.hover_Сard_4_10.hide()
    def start_timer_Сard_4_10_FO(self, event):
        self.timer_Сard_4_10_FO.start(2000)  # 2000 milliseconds = 2 seconds
    def WINdow_Сard_4_10_FO(self):
        self.timer_Сard_4_10_FO.stop()
        self.hover_Сard_4_10_FO.move(self.UG.ui.card_5_btn_5.mapToGlobal(self.UG.ui.card_5_btn_5.rect().bottomLeft()))
        self.hover_Сard_4_10_FO.show()
    def WINdow_hide_Сard_4_10_FO(self, event):
        self.timer_Сard_4_10_FO.stop()
        self.hover_Сard_4_10_FO.hide()
    def start_timer_Сard_1_40(self, event):
        self.timer_Сard_1_40.start(2000)  # 2000 milliseconds = 2 seconds
    def WINdow_Сard_1_40(self):
        self.timer_Сard_1_40.stop()
        self.hover_Сard_1_40.move(self.UG.ui.card_6_btn_6.mapToGlobal(self.UG.ui.card_6_btn_6.rect().bottomLeft()))
        self.hover_Сard_1_40.show()
    def WINdow_hide_Сard_1_40(self, event):
        self.timer_Сard_1_40.stop()
        self.hover_Сard_1_40.hide()
    def start_timer_SFP_RJ_45_1G(self, event):
        self.timer_SFP_RJ_45_1G.start(2000)
    def WINdow_SFP_RJ_45_1G(self):
        self.timer_SFP_RJ_45_1G.stop()
        self.hover_SFP_RJ_45_1G.move(self.UG.ui.card_7_btn_7.mapToGlobal(self.UG.ui.card_7_btn_7.rect().bottomLeft()))
        self.hover_SFP_RJ_45_1G.show()
    def WINdow_hide_SFP_RJ_45_1G(self, event):
        self.timer_SFP_RJ_45_1G.stop()
        self.hover_SFP_RJ_45_1G.hide()
    def start_timer_SFP_10G(self, event):
        self.timer_SFP_10G.start(2000)
    def WINdow_SFP_10G(self):
        self.timer_SFP_10G.stop()
        self.hover_SFP_10G.move(self.UG.ui.card_8_btn_8.mapToGlobal(self.UG.ui.card_8_btn_8.rect().bottomLeft()))
        self.hover_SFP_10G.show()
    def WINdow_hide_SFP_10G(self, event):
        self.timer_SFP_10G.stop()
        self.hover_SFP_10G.hide()
    def start_timer_SFP_10Gb_500m(self, event):
        self.timer_SFP_10Gb_500m.start(2000)
    def WINdow_SFP_10Gb_500m(self):
        self.timer_SFP_10Gb_500m.stop()
        self.hover_SFP_10Gb_500m.move(self.UG.ui.card_9_btn_9.mapToGlobal(self.UG.ui.card_9_btn_9.rect().bottomLeft()))
        self.hover_SFP_10Gb_500m.show()
    def WINdow_hide_SFP_10Gb_500m(self, event):
        self.timer_SFP_10Gb_500m.stop()
        self.hover_SFP_10Gb_500m.hide()
    def start_timer_SFP_10Gb_10km(self, event):
        self.timer_SFP_10Gb_10km.start(2000)
    def WINdow_SFP_10Gb_10km(self):
        self.timer_SFP_10Gb_10km.stop()
        self.hover_SFP_10Gb_10km.move(self.UG.ui.card_10_btn_10.mapToGlobal(self.UG.ui.card_10_btn_10.rect().bottomLeft()))
        self.hover_SFP_10Gb_10km.show()
    def WINdow_hide_SFP_10Gb_10km(self, event):
        self.timer_SFP_10Gb_10km.stop()
        self.hover_SFP_10Gb_10km.hide()
    def start_timer_SFP_SX_1_25(self, event):
        self.timer_SFP_SX_1_25.start(2000)
    def WINdow_SFP_SX_1_25(self):
        self.timer_SFP_SX_1_25.stop()
        self.hover_SFP_SX_1_25.move(self.UG.ui.card_11_btn_11.mapToGlobal(self.UG.ui.card_11_btn_11.rect().bottomLeft()))
        self.hover_SFP_SX_1_25.show()
    def WINdow_hide_SFP_SX_1_25(self, event):
        self.timer_SFP_SX_1_25.stop()
        self.hover_SFP_SX_1_25.hide()
    ##############################################################################
    ###event clicked on treeView###
    def onItemClicked(self, item):
        self.selected_name_in_view=(item.text(0))
    ##############################################################################
    ###event checkbox###
    def on_check_yes_changed(self, state):
        if state == 2:  # 2 означает, что чекбокс включен
            self.UG.ui.check_no.setChecked(False)
            self.fstec_status = "Да"
        else:
            self.fstec_status = None
    def on_check_no_changed(self, state):
        if state == 2:  # 2 означает, что чекбокс включен
            self.UG.ui.check_yes.setChecked(False)
            self.fstec_status = "Нет"
        else:
            self.fstec_status = None
    def standart_changed(self, state):
        if state == 2:  # 2 означает, что чекбокс включен
            self.UG.ui.premium_checkBox.setChecked(False)
            self.TP_status = "Standart"
        else:
            self.TP_status = None
    def premium_changed(self, state):
        if state == 2:  # 2 означает, что чекбокс включен
            self.UG.ui.standart_checkBox.setChecked(False)
            self.TP_status = "Premium"
        else:
            self.TP_status = None
    ##############################################################################
    ###event spinbox###
    def spinbox_yearBox_changed(self, value):
        self.value_year_moduls = value  # Записываем значение в переменную
    def spinbox_yearBox_tp_changed(self, value):
        self.value_year_tp = value  
    ##############################################################################
    ###event text comment###
    def on_text_coment_changed(self):
        self.text_coment = self.UG.ui.comentEdit.toPlainText()
                